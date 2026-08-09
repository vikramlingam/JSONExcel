"""Read JSON values stored in Excel cells into the normal record pipeline."""

import json
import re
from collections.abc import Mapping
from typing import Any

from openpyxl import load_workbook

from .exceptions import ConfigurationError, ConversionError, InvalidInputError


def read_json_column(source: Any, *, sheet: str | None, json_column: str, separator: str = ".", arrays: str = "json", include_source_row: bool = False, include_columns: list[str] | None = None, header_order: str = "first-seen", preferred_columns: list[str] | None = None, errors: str = "report", clean_json: bool = False) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    if arrays not in {"json", "join"}:
        raise ConfigurationError("arrays must be 'json' or 'join'.")
    if errors not in {"report", "skip", "raise"}:
        raise ConfigurationError("errors must be 'report', 'skip', or 'raise'.")
    if header_order not in {"first-seen", "alphabetical"}:
        raise ConfigurationError("header_order must be 'first-seen' or 'alphabetical'.")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = _select_sheet(workbook, sheet)
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            raise InvalidInputError("The source workbook is empty.")
        headers = _headers(raw_headers)
        if json_column not in headers:
            raise InvalidInputError(f"JSON column '{json_column}' was not found in worksheet '{worksheet.title}'.")
        if include_columns:
            missing = [name for name in include_columns if name not in headers]
            if missing:
                raise InvalidInputError(f"Source columns were not found: {', '.join(missing)}")
        json_index = headers.index(json_column)
        include_indexes = [(name, headers.index(name)) for name in include_columns or []]
        records: list[dict[str, Any]] = []
        errors_found: list[dict[str, Any]] = []
        for excel_row, values in enumerate(rows, start=2):
            raw_value = values[json_index] if json_index < len(values) else None
            if raw_value is None or not str(raw_value).strip():
                continue
            try:
                parsed = _parse_cell(str(raw_value), clean_json=clean_json)
                objects = parsed if isinstance(parsed, list) else [parsed]
                if isinstance(parsed, list) and parsed and all(not isinstance(item, (Mapping, list)) for item in parsed):
                    objects = [{json_column: parsed}]
                if not all(isinstance(item, Mapping) for item in objects):
                    raise ValueError("JSON value must be an object or an array of objects")
                for item in objects:
                    record = {name: values[index] if index < len(values) else None for name, index in include_indexes}
                    if include_source_row:
                        record = {"_source_row": excel_row, **record}
                    record.update(_flatten_json(dict(item), separator=separator, arrays=arrays))
                    records.append(record)
            except (json.JSONDecodeError, ValueError) as exc:
                message = _error_message(exc)
                if errors == "raise":
                    raise ConversionError(f"Invalid JSON in worksheet '{worksheet.title}', row {excel_row}: {message}") from exc
                if errors == "report":
                    errors_found.append({"source_row": excel_row, "error": message, "raw_value": str(raw_value)})
        field_order = _field_order(records, header_order=header_order, preferred_columns=preferred_columns)
        return records, errors_found, {"headers": headers, "fields": field_order, "worksheet": worksheet.title}
    finally:
        workbook.close()


def read_json_columns(source: Any, *, sheet: str | None, json_columns: list[str], separator: str = ".", arrays: str = "json", include_source_row: bool = False, include_columns: list[str] | None = None, header_order: str = "first-seen", preferred_columns: list[str] | None = None, errors: str = "report", clean_json: bool = False) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Read multiple JSON columns and prefix discovered fields by column name."""
    if len(json_columns) < 2:
        raise ConfigurationError("json_columns requires at least two column names.")
    if len(set(json_columns)) != len(json_columns):
        raise ConfigurationError("json_columns must not contain duplicates.")
    if arrays not in {"json", "join"}:
        raise ConfigurationError("arrays must be 'json' or 'join'.")
    if errors not in {"report", "skip", "raise"}:
        raise ConfigurationError("errors must be 'report', 'skip', or 'raise'.")
    if header_order not in {"first-seen", "alphabetical"}:
        raise ConfigurationError("header_order must be 'first-seen' or 'alphabetical'.")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = _select_sheet(workbook, sheet)
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            raise InvalidInputError("The source workbook is empty.")
        headers = _headers(raw_headers)
        missing_json = [name for name in json_columns if name not in headers]
        if missing_json:
            raise InvalidInputError(f"JSON columns were not found: {', '.join(missing_json)}")
        if include_columns:
            missing = [name for name in include_columns if name not in headers]
            if missing:
                raise InvalidInputError(f"Source columns were not found: {', '.join(missing)}")
        indexes = {name: headers.index(name) for name in headers}
        records: list[dict[str, Any]] = []
        errors_found: list[dict[str, Any]] = []
        for excel_row, values in enumerate(rows, start=2):
            parsed_columns: list[tuple[str, list[dict[str, Any]]]] = []
            row_had_value = False
            valid_json_count = 0
            for column in json_columns:
                raw_value = values[indexes[column]] if indexes[column] < len(values) else None
                if raw_value is None or not str(raw_value).strip():
                    parsed_columns.append((column, [{}]))
                    continue
                row_had_value = True
                try:
                    parsed = _parse_cell(str(raw_value), clean_json=clean_json)
                    parsed_columns.append((column, _as_objects(parsed, column)))
                    valid_json_count += 1
                except (json.JSONDecodeError, ValueError) as exc:
                    message = _error_message(exc)
                    if errors == "raise":
                        raise ConversionError(f"Invalid JSON in worksheet '{worksheet.title}', row {excel_row}, column '{column}': {message}") from exc
                    if errors == "report":
                        errors_found.append({"source_row": excel_row, "json_column": column, "error": message, "raw_value": str(raw_value)})
                    parsed_columns.append((column, [{}]))
            if not row_had_value or valid_json_count == 0:
                continue
            combinations: list[dict[str, Any]] = [{}]
            for column, objects in parsed_columns:
                combinations = [{**base, **_prefix_record(item, column, separator, arrays)} for base in combinations for item in objects]
            for combination in combinations:
                record = {name: values[indexes[name]] if indexes[name] < len(values) else None for name in include_columns or []}
                if include_source_row:
                    record = {"_source_row": excel_row, **record}
                record.update(combination)
                records.append(record)
        field_order = _field_order(records, header_order=header_order, preferred_columns=preferred_columns)
        return records, errors_found, {"headers": headers, "fields": field_order, "worksheet": worksheet.title}
    finally:
        workbook.close()


def inspect_json_column(source: Any, *, sheet: str | None = None, json_column: str = "JSON", separator: str = ".", arrays: str = "json", clean_json: bool = False) -> dict[str, Any]:
    records, errors, plan = read_json_column(source, sheet=sheet, json_column=json_column, separator=separator, arrays=arrays, errors="report", clean_json=clean_json)
    present: dict[str, int] = {}
    for record in records:
        for field in record:
            present[field] = present.get(field, 0) + 1
    total = len(records)
    return {"records": total, "valid_records": total, "invalid_records": len(errors), "fields": {field: {"present": count, "missing": total - count, "coverage": count / total if total else 0.0} for field, count in present.items()}, "worksheet": plan["worksheet"]}


def _select_sheet(workbook: Any, sheet: str | None) -> Any:
    if sheet:
        if sheet not in workbook.sheetnames:
            raise InvalidInputError(f"Worksheet '{sheet}' was not found. Available worksheets: {', '.join(workbook.sheetnames)}")
        return workbook[sheet]
    if len(workbook.sheetnames) != 1:
        raise InvalidInputError("The workbook has multiple worksheets; specify sheet='SheetName'.")
    return workbook[workbook.sheetnames[0]]


def _headers(values: tuple[Any, ...]) -> list[str]:
    headers = [str(value) if value is not None else "" for value in values]
    if len(headers) != len(set(headers)):
        duplicates = sorted({header for header in headers if header and headers.count(header) > 1})
        raise InvalidInputError(f"Worksheet has duplicate column headers: {', '.join(duplicates)}")
    return headers


def _parse_cell(value: str, *, clean_json: bool) -> Any:
    text = value.strip()
    if clean_json:
        match = re.fullmatch(r"```(?:json)?\s*(.*?)\s*```", text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            text = match.group(1).strip()
    return json.loads(text)


def _as_objects(parsed: Any, column: str) -> list[dict[str, Any]]:
    if isinstance(parsed, list) and parsed and all(not isinstance(item, (Mapping, list)) for item in parsed):
        return [{column: parsed}]
    objects = parsed if isinstance(parsed, list) else [parsed]
    if not all(isinstance(item, Mapping) for item in objects):
        raise ValueError("JSON value must be an object or an array of objects")
    return [dict(item) for item in objects]


def _prefix_record(record: dict[str, Any], column: str, separator: str, arrays: str = "json") -> dict[str, Any]:
    return {f"{column}{separator}{key}": value for key, value in _flatten_json(record, separator=separator, arrays=arrays).items()}


def _flatten_json(record: dict[str, Any], *, separator: str, arrays: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    def visit(prefix: str, value: Any) -> None:
        if isinstance(value, Mapping):
            for key, child in value.items():
                visit(f"{prefix}{separator}{key}" if prefix else str(key), child)
        elif isinstance(value, list):
            if arrays == "join" and all(not isinstance(item, (Mapping, list)) for item in value):
                result[prefix] = ", ".join(str(item) for item in value)
            else:
                result[prefix] = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        else:
            result[prefix] = value
    visit("", record)
    return result


def _field_order(records: list[dict[str, Any]], *, header_order: str, preferred_columns: list[str] | None) -> list[str]:
    discovered: list[str] = []
    seen: set[str] = set()
    for record in records:
        for field in record:
            if field not in seen:
                seen.add(field)
                discovered.append(field)
    preferred = [field for field in preferred_columns or [] if field in seen]
    remaining = [field for field in discovered if field not in preferred]
    return preferred + (sorted(remaining) if header_order == "alphabetical" else remaining)


def _error_message(exc: Exception) -> str:
    if isinstance(exc, json.JSONDecodeError):
        return f"Invalid JSON: {exc.msg}"
    return str(exc)
