"""Tabular workbook reading and explicit embedded-source restoration."""

from typing import Any
import hashlib
import json

from openpyxl import load_workbook

from .exceptions import ConversionError, InvalidInputError
from .writer import METADATA_STORAGE_VERSION


def from_excel(source: Any, *, restore_source: bool = False) -> Any:
    """Read visible tabular data or retrieve an explicitly embedded source snapshot.

    ``restore_source=True`` returns the saved JSON-compatible source value. It does
    not reconstruct edits made to visible worksheet cells.
    """
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if restore_source:
            stored = _read_metadata(workbook)
            if stored is None:
                raise InvalidInputError("Workbook does not contain an embedded jsonexcel source snapshot.")
            if "source_snapshot" not in stored:
                raise ConversionError("jsonexcel metadata does not contain a source snapshot.")
            return stored["source_snapshot"]
        names = [name for name in workbook.sheetnames if not name.startswith("__jsonexcel") and name not in {"Summary", "Errors", "Dashboard"}]
        return _read_sheet(workbook[names[0]]) if names else []
    finally:
        workbook.close()


def metadata(mode: str, tables: list[str], *, source_snapshot: Any = None, relationships: list[dict[str, str]] | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {"payload_version": 1, "kind": "source_snapshot", "mode": mode, "tables": tables}
    if relationships:
        result["relationships"] = relationships
    if source_snapshot is not None:
        result["source_snapshot"] = source_snapshot
    return result


def _read_metadata(workbook: Any) -> dict[str, Any] | None:
    name = "__jsonexcel_metadata"
    if name not in workbook.sheetnames:
        return None
    rows = list(workbook[name].iter_rows(values_only=True))
    if not rows:
        raise ConversionError("jsonexcel metadata worksheet is empty.")
    entries: dict[str, Any] = {}
    for row in rows:
        if not row or row[0] is None:
            continue
        label = str(row[0])
        if label in entries:
            raise ConversionError(f"jsonexcel metadata contains duplicate entry '{label}'.")
        entries[label] = row[1] if len(row) > 1 else None
    try:
        version = int(entries["metadata_version"])
        chunk_count = int(entries["chunk_count"])
        expected_hash = str(entries["sha256"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ConversionError("jsonexcel metadata header is missing or malformed.") from exc
    if version != METADATA_STORAGE_VERSION:
        raise ConversionError(f"Unsupported jsonexcel metadata storage version: {version}.")
    if chunk_count < 1:
        raise ConversionError("jsonexcel metadata chunk count must be at least one.")
    chunks: list[str] = []
    for index in range(1, chunk_count + 1):
        label = f"chunk_{index:06d}"
        value = entries.get(label)
        if not isinstance(value, str):
            raise ConversionError(f"jsonexcel metadata is missing or has malformed chunk '{label}'.")
        chunks.append(value)
    actual_chunk_labels = {label for label in entries if label.startswith("chunk_") and label[6:].isdigit()}
    expected_chunk_labels = {f"chunk_{index:06d}" for index in range(1, chunk_count + 1)}
    if actual_chunk_labels != expected_chunk_labels:
        raise ConversionError("jsonexcel metadata chunk sequence is incomplete or contains unexpected chunks.")
    serialized = "".join(chunks)
    actual_hash = hashlib.sha256(serialized.encode("utf-8")).hexdigest()
    if actual_hash != expected_hash:
        raise ConversionError("jsonexcel metadata checksum validation failed.")
    try:
        value = json.loads(serialized)
    except json.JSONDecodeError as exc:
        raise ConversionError("jsonexcel metadata JSON is malformed.") from exc
    if not isinstance(value, dict):
        raise ConversionError("jsonexcel metadata payload must be a JSON object.")
    return value


def _read_sheet(sheet: Any) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [{header: row[index] for index, header in enumerate(headers) if index < len(row) and row[index] is not None} for row in rows[1:]]
